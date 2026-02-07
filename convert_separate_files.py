import openpyxl
import json

# Deyimler dosyasını oku
print("Deyimler dosyası okunuyor...")
wb_deyimler = openpyxl.load_workbook('deyimler_duzeltilmis.xlsx')
ws_deyimler = wb_deyimler.active

deyimler = []
row_num = 0
for row in ws_deyimler.iter_rows(min_row=1, values_only=True):
    row_num += 1
    if not row[0]:  # Boş satırları atla
        continue
    
    text = str(row[0]).strip() if row[0] else ""
    meaning = str(row[1]).strip() if len(row) > 1 and row[1] else ""
    
    # Başlık satırını atla
    if text.lower() in ['ifade', 'deyim', 'atasözü', 'text'] or meaning.lower() in ['anlam', 'meaning']:
        continue
    
    deyim = {
        "id": len(deyimler) + 1,
        "text": text,
        "meaning": meaning
    }
    deyimler.append(deyim)

# Deyimler JSON'a kaydet
with open('deyimler.json', 'w', encoding='utf-8') as f:
    json.dump({"deyimler": deyimler}, f, ensure_ascii=False, indent=2)

print(f"✅ deyimler.json oluşturuldu - {len(deyimler)} adet deyim")

# Atasözleri dosyasını oku
print("\nAtasözleri dosyası okunuyor...")
wb_atasozleri = openpyxl.load_workbook('atasozleri_duzeltilmis.xlsx')
ws_atasozleri = wb_atasozleri.active

atasozleri = []
row_num = 0
for row in ws_atasozleri.iter_rows(min_row=1, values_only=True):
    row_num += 1
    if not row[0]:  # Boş satırları atla
        continue
    
    text = str(row[0]).strip() if row[0] else ""
    meaning = str(row[1]).strip() if len(row) > 1 and row[1] else ""
    
    # Başlık satırını atla
    if text.lower() in ['ifade', 'deyim', 'atasözü', 'text'] or meaning.lower() in ['anlam', 'meaning']:
        continue
    
    atasozu = {
        "id": len(atasozleri) + 1,
        "text": text,
        "meaning": meaning
    }
    atasozleri.append(atasozu)

# Atasözleri JSON'a kaydet
with open('atasozleri.json', 'w', encoding='utf-8') as f:
    json.dump({"atasozleri": atasozleri}, f, ensure_ascii=False, indent=2)

print(f"✅ atasozleri.json oluşturuldu - {len(atasozleri)} adet atasözü")

print("\n🎉 Tüm dosyalar başarıyla oluşturuldu!")
